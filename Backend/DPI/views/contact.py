from django.http import JsonResponse
from openpyxl import Workbook, load_workbook
import os
from django.views.decorators.csrf import csrf_exempt
from threading import Lock

# Initialize a thread lock for thread-safe file access
file_lock = Lock()

# Path to the Excel file
excel_file_path = 'contact_data.xlsx'

# Create the Excel file if it does not exist
if not os.path.exists(excel_file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Email", "Object", "Message"])  # Header row
    wb.save(excel_file_path)

@csrf_exempt
def save_contact_form(request):
    """
    Save contact form data (email, object, and message) to an Excel file.

    This view accepts POST requests with form data and appends the data 
    to an existing Excel file. If the file does not exist, it creates one.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        JsonResponse: A JSON response indicating success or failure.

    Methods:
        POST: 
            - Accepts form data with fields: `email`, `object`, `message`.
            - Appends the data to the Excel file.
        Any other method:
            - Returns a 400 status with an error message.
    """
    if request.method == "POST":
        # Retrieve form data
        email = request.POST.get("email")
        object_ = request.POST.get("object")
        message = request.POST.get("message")

        # Validate required fields
        if not email or not object_ or not message:
            return JsonResponse({"message": "Missing required fields"}, status=400)

        try:
            with file_lock:
                # Open the existing Excel file and append data
                wb = load_workbook(excel_file_path)
                ws = wb.active
                ws.append([email, object_, message])  # Add data as a new row
                wb.save(excel_file_path)

            return JsonResponse({"message": "Data saved successfully"}, status=200)

        except Exception as e:
            return JsonResponse({"message": f"Error saving data: {str(e)}"}, status=500)

    else:
        return JsonResponse({"message": "Invalid request"}, status=400)